#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FANGIO TELECOM - Versión Simple para Heroku
Sistema de Llenado Automático para equipo distribuido
"""

import os
import sys
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import tempfile
import shutil

# Configuración de la aplicación
app = Flask(__name__)
app.secret_key = 'fangio_telecom_2024'

# Configuración de rutas
base_dir = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(base_dir, 'uploads')
OUTPUT_FOLDER = os.path.join(base_dir, 'Temp')

# Crear carpetas si no existen
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Configuración de archivos permitidos
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'xlsx', 'kmz'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>FANGIO TELECOM</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
            .container { max-width: 800px; margin: 0 auto; background: rgba(255,255,255,0.1); padding: 30px; border-radius: 15px; }
            h1 { text-align: center; color: #fff; text-shadow: 2px 2px 4px rgba(0,0,0,0.3); }
            .btn { display: inline-block; padding: 15px 30px; margin: 10px; background: #4CAF50; color: white; text-decoration: none; border-radius: 8px; font-size: 18px; }
            .btn:hover { background: #45a049; }
            .status { background: rgba(255,255,255,0.2); padding: 20px; border-radius: 10px; margin: 20px 0; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 FANGIO TELECOM</h1>
            <h2>Sistema de Llenado Automático</h2>
            
            <div class="status">
                <h3>✅ Estado: Conectado</h3>
                <p>Servidor funcionando correctamente en Heroku</p>
            </div>
            
            <h3>Funcionalidades Disponibles:</h3>
            <a href="/site_survey" class="btn">📋 Site Survey</a>
            <a href="/diseno_solucion" class="btn">🔧 Diseño de Solución</a>
            <a href="/status" class="btn">📊 Estado del Servidor</a>
            
            <h3>Para tu equipo:</h3>
            <p>Comparte esta URL con tu equipo en Guadalajara:</p>
            <p><strong>https://fangio-telecom-1579cec4f184.herokuapp.com</strong></p>
        </div>
    </body>
    </html>
    '''

@app.route('/site_survey', methods=['GET', 'POST'])
def site_survey():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            archivos = request.files.getlist('archivos')
            
            # Guardar archivos subidos
            archivos_guardados = []
            for archivo in archivos:
                if archivo and allowed_file(archivo.filename):
                    filename = secure_filename(archivo.filename)
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    archivo.save(filepath)
                    archivos_guardados.append(filepath)
            
            # Crear archivo Excel
            output_filename = f"site_survey_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Site Survey"
            
            # Llenar datos
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Agregar archivos
            ws[f'A{row}'] = "Archivos subidos:"
            row += 1
            for archivo in archivos_guardados:
                ws[f'A{row}'] = os.path.basename(archivo)
                row += 1
            
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name=output_filename)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Site Survey - FANGIO TELECOM</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
            .container { max-width: 600px; margin: 0 auto; background: rgba(255,255,255,0.1); padding: 30px; border-radius: 15px; }
            input, textarea, select { width: 100%; padding: 10px; margin: 5px 0; border: none; border-radius: 5px; }
            .btn { padding: 15px 30px; background: #4CAF50; color: white; border: none; border-radius: 8px; font-size: 16px; cursor: pointer; }
            .btn:hover { background: #45a049; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📋 Site Survey</h1>
            <form method="POST" enctype="multipart/form-data">
                <h3>Información del Sitio:</h3>
                <input type="text" name="sitio" placeholder="Nombre del sitio" required>
                <input type="text" name="direccion" placeholder="Dirección" required>
                <input type="text" name="coordenadas" placeholder="Coordenadas GPS">
                
                <h3>Información Técnica:</h3>
                <input type="text" name="tipo_antena" placeholder="Tipo de antena">
                <input type="text" name="altura" placeholder="Altura de instalación">
                <input type="text" name="potencia" placeholder="Potencia de transmisión">
                
                <h3>Archivos:</h3>
                <input type="file" name="archivos" multiple accept=".pdf,.png,.jpg,.jpeg,.docx,.xlsx,.kmz">
                
                <br><br>
                <button type="submit" class="btn">📥 Generar Site Survey</button>
            </form>
            <br>
            <a href="/" style="color: white;">← Volver al inicio</a>
        </div>
    </body>
    </html>
    '''

@app.route('/diseno_solucion', methods=['GET', 'POST'])
def diseno_solucion():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            archivos = request.files.getlist('archivos')
            
            # Guardar archivos subidos
            archivos_guardados = []
            for archivo in archivos:
                if archivo and allowed_file(archivo.filename):
                    filename = secure_filename(archivo.filename)
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    archivo.save(filepath)
                    archivos_guardados.append(filepath)
            
            # Crear archivo Excel
            output_filename = f"diseno_solucion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Diseño de Solución"
            
            # Llenar datos
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Agregar archivos
            ws[f'A{row}'] = "Archivos subidos:"
            row += 1
            for archivo in archivos_guardados:
                ws[f'A{row}'] = os.path.basename(archivo)
                row += 1
            
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name=output_filename)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Diseño de Solución - FANGIO TELECOM</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
            .container { max-width: 600px; margin: 0 auto; background: rgba(255,255,255,0.1); padding: 30px; border-radius: 15px; }
            input, textarea, select { width: 100%; padding: 10px; margin: 5px 0; border: none; border-radius: 5px; }
            .btn { padding: 15px 30px; background: #4CAF50; color: white; border: none; border-radius: 8px; font-size: 16px; cursor: pointer; }
            .btn:hover { background: #45a049; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔧 Diseño de Solución</h1>
            <form method="POST" enctype="multipart/form-data">
                <h3>Información del Proyecto:</h3>
                <input type="text" name="proyecto" placeholder="Nombre del proyecto" required>
                <input type="text" name="cliente" placeholder="Cliente" required>
                <textarea name="descripcion" placeholder="Descripción del proyecto" rows="4"></textarea>
                
                <h3>Especificaciones Técnicas:</h3>
                <input type="text" name="equipos" placeholder="Equipos a utilizar">
                <input type="text" name="materiales" placeholder="Materiales necesarios">
                <input type="text" name="tiempo_estimado" placeholder="Tiempo estimado de instalación">
                
                <h3>Archivos:</h3>
                <input type="file" name="archivos" multiple accept=".pdf,.png,.jpg,.jpeg,.docx,.xlsx,.kmz">
                
                <br><br>
                <button type="submit" class="btn">📥 Generar Diseño de Solución</button>
            </form>
            <br>
            <a href="/" style="color: white;">← Volver al inicio</a>
        </div>
    </body>
    </html>
    '''

@app.route('/status')
def status():
    return jsonify({
        'status': 'online',
        'service': 'FANGIO TELECOM',
        'version': '2.0-simple',
        'timestamp': datetime.now().isoformat(),
        'features': ['Site Survey', 'Diseño de Solución', 'Subida de archivos'],
        'url': 'https://fangio-telecom-1579cec4f184.herokuapp.com'
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False) 