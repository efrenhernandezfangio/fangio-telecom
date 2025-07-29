import os
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask, request, send_file, render_template_string, redirect, url_for, after_this_request
from werkzeug.utils import secure_filename
import sys
import re
import dataframe_image as dfi
import matplotlib.pyplot as plt
import textwrap
import unicodedata
import glob

def normaliza_na(valor):
    if isinstance(valor, str) and valor.strip().lower() == "n/a":
        return "N/A"
    elif pd.isna(valor):
        return "N/A"
    elif valor == "" or (isinstance(valor, str) and valor.strip() == ""):
        return "N/A"
    return valor

# Obtener el directorio base de la aplicación
if getattr(sys, 'frozen', False):
    # Si es un ejecutable compilado
    base_dir = os.path.dirname(sys.executable)
else:
    # Si es código Python normal
    base_dir = os.path.dirname(os.path.abspath(__file__))

print(f"Directorio base: {base_dir}")
print("Archivos en el directorio:", os.listdir(base_dir))

# Buscar el archivo en múltiples ubicaciones
llenado_paths = [
    os.path.join(base_dir, 'llenado-automatico.html'),
    os.path.join(base_dir, 'static', 'llenado-automatico.html'),
    os.path.join(base_dir, 'templates', 'llenado-automatico.html'),
    'llenado-automatico.html'  # Directorio actual como fallback
]

html_form = None
for path in llenado_paths:
    try:
        with open(path, encoding='utf-8') as f:
            html_form = f.read()
        print(f"Archivo llenado-automatico.html cargado desde: {path}")
        break
    except Exception as e:
        print(f"No se pudo cargar desde {path}: {e}")
        continue

if html_form is None:
    print("ERROR: No se pudo cargar llenado-automatico.html desde ninguna ubicación")
    print("Ubicaciones probadas:")
    for path in llenado_paths:
        print(f"  - {path}")
    print("Asegúrate de que el archivo existe en una de estas ubicaciones")
    input("Presiona Enter para salir...")
    sys.exit(1)

app = Flask(__name__)

# Usar ruta relativa para que funcione en cualquier computadora
import os
base_dir = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(base_dir, 'site_survey')
GOOGLE_SHEETS_CSV_URL = 'https://docs.google.com/spreadsheets/d/1sfOY1Y3dNVCOT8zyCMzpgARv-R_jRE-S/export?format=csv'

@app.route('/site_survey_checkboxes', methods=['GET'])
def site_survey_checkboxes():
    import pandas as pd
    user_id = request.args.get('user_id', '')
    fila_idx = request.args.get('fila_idx', '')
    chk_urbana = chk_suburbana = chk_rural = chk_ejidal = chk_pueblo_magico = False
    if fila_idx:
        try:
            df_db = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
            row = df_db.loc[int(fila_idx)]
            tipo_zona_original = row.get('Tipo de Zona', '')
            tipo_zona = normaliza_texto(tipo_zona_original)
            chk_urbana = 'urbana' in tipo_zona
            chk_suburbana = 'suburbana' in tipo_zona or 'suburbana' in tipo_zona or 'suburbana' in tipo_zona.replace('sub', '')
            chk_rural = 'rural' in tipo_zona
            chk_ejidal = 'ejidal' in tipo_zona
            chk_pueblo_magico = 'pueblomagico' in tipo_zona
        except Exception as e:
            print(f"Error leyendo base de datos: {e}")
    return render_template(
        'site_survey_checkboxes.html',
        chk_urbana=chk_urbana,
        chk_suburbana=chk_suburbana,
        chk_rural=chk_rural,
        chk_ejidal=chk_ejidal,
        chk_pueblo_magico=chk_pueblo_magico
    )

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template_string(html_form)

@app.route('/diseno_solucion', methods=['GET', 'POST'])
def diseno_solucion():
    if request.method == 'POST':
        # Procesar formulario de diseño de solución
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            
            # Crear archivo Excel con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Diseño de Solución"
            
            # Llenar datos en el Excel
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Guardar archivo
            output_path = os.path.join(base_dir, 'site_survey', 'diseno_solucion.xlsx')
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name='diseno_solucion.xlsx')
            
        except Exception as e:
            return f"Error: {str(e)}", 500
    
    # Mostrar formulario
    return render_template_string(html_form)

@app.route('/site_survey', methods=['GET'])
def site_survey():
    try:
        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear tabla HTML
        tabla_html = df.to_html(classes='table table-striped', index=False)
        
        return render_template_string(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Site Survey - FANGIO TELECOM</title>
            <style>
                body {{
                    background-image: url('/static/images/earth-background.jpg');
                    background-size: cover;
                    background-attachment: fixed;
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: white;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: rgba(0, 0, 0, 0.8);
                    padding: 20px;
                    border-radius: 10px;
                }}
                .logo {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .logo img {{
                    max-width: 200px;
                    height: auto;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    background: rgba(255, 255, 255, 0.9);
                    color: black;
                }}
                th, td {{
                    padding: 8px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #4CAF50;
                    color: white;
                }}
                .btn {{
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                }}
                .btn:hover {{
                    background-color: #45a049;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
                </div>
                <h1>Base de Datos - Site Survey</h1>
                <a href="/" class="btn">← Volver al Inicio</a>
                {tabla_html}
            </div>
        </body>
        </html>
        """)
        
    except Exception as e:
        return f"Error al cargar datos: {str(e)}", 500

@app.route('/descargar_site_survey')
def descargar_site_survey():
    try:
        def limpiar_nombre_archivo(nombre):
            # Eliminar caracteres problemáticos para nombres de archivo
            caracteres_invalidos = r'[<>:"/\\|?*]'
            nombre_limpio = re.sub(caracteres_invalidos, '_', str(nombre))
            return nombre_limpio

        @after_this_request
        def eliminar_archivos_temporales(response):
            try:
                # Limpiar archivos temporales si es necesario
                pass
            except Exception as e:
                print(f"Error limpiando archivos: {e}")
            return response

        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear archivo Excel con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Site Survey"
        
        # Llenar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar archivo
        output_path = os.path.join(base_dir, 'site_survey', 'site_survey_completo.xlsx')
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name='site_survey_completo.xlsx')
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/reporte_planeacion')
def reporte_planeacion():
    try:
        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear reporte de planeación
        reporte_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Reporte de Planeación - FANGIO TELECOM</title>
            <style>
                body {{
                    background-image: url('/static/images/earth-background.jpg');
                    background-size: cover;
                    background-attachment: fixed;
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: white;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: rgba(0, 0, 0, 0.8);
                    padding: 20px;
                    border-radius: 10px;
                }}
                .logo {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .logo img {{
                    max-width: 200px;
                    height: auto;
                }}
                .btn {{
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                }}
                .btn:hover {{
                    background-color: #45a049;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
                </div>
                <h1>Reporte de Planeación</h1>
                <a href="/" class="btn">← Volver al Inicio</a>
                <p>Total de registros: {len(df)}</p>
                <p>Reporte generado el: {time.strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """
        
        return render_template_string(reporte_html)
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/formulario_archivos', methods=['GET'])
def formulario_archivos():
    return render_template_string(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Formulario de Archivos - FANGIO TELECOM</title>
        <style>
            body {{
                background-image: url('/static/images/earth-background.jpg');
                background-size: cover;
                background-attachment: fixed;
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: white;
            }}
            .container {{
                max-width: 800px;
                margin: 0 auto;
                background: rgba(0, 0, 0, 0.8);
                padding: 20px;
                border-radius: 10px;
            }}
            .logo {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .logo img {{
                max-width: 200px;
                height: auto;
            }}
            .form-group {{
                margin-bottom: 15px;
            }}
            label {{
                display: block;
                margin-bottom: 5px;
                font-weight: bold;
            }}
            input[type="file"], input[type="text"] {{
                width: 100%;
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                box-sizing: border-box;
            }}
            .btn {{
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
                margin: 5px;
            }}
            .btn:hover {{
                background-color: #45a049;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logo">
                <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
            </div>
            <h1>Formulario de Archivos</h1>
            <a href="/" class="btn">← Volver al Inicio</a>
            <form action="/procesar" method="post" enctype="multipart/form-data">
                <div class="form-group">
                    <label for="archivo">Seleccionar archivo:</label>
                    <input type="file" id="archivo" name="archivo" required>
                </div>
                <div class="form-group">
                    <label for="descripcion">Descripción:</label>
                    <input type="text" id="descripcion" name="descripcion" required>
                </div>
                <button type="submit" class="btn">Procesar</button>
            </form>
        </div>
    </body>
    </html>
    """)

@app.route('/seleccion', methods=['POST'])
def seleccion():
    try:
        # Obtener datos del formulario
        datos = request.form.to_dict()
        
        # Procesar selección
        return render_template_string(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Selección Procesada - FANGIO TELECOM</title>
            <style>
                body {{
                    background-image: url('/static/images/earth-background.jpg');
                    background-size: cover;
                    background-attachment: fixed;
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: white;
                }}
                .container {{
                    max-width: 800px;
                    margin: 0 auto;
                    background: rgba(0, 0, 0, 0.8);
                    padding: 20px;
                    border-radius: 10px;
                }}
                .logo {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .logo img {{
                    max-width: 200px;
                    height: auto;
                }}
                .btn {{
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                }}
                .btn:hover {{
                    background-color: #45a049;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
                </div>
                <h1>Selección Procesada</h1>
                <a href="/" class="btn">← Volver al Inicio</a>
                <p>Datos recibidos: {datos}</p>
            </div>
        </body>
        </html>
        """)
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/procesar', methods=['POST'])
def procesar():
    try:
        # Procesar archivo subido
        if 'archivo' not in request.files:
            return 'No se seleccionó ningún archivo', 400
        
        archivo = request.files['archivo']
        descripcion = request.form.get('descripcion', '')
        
        if archivo.filename == '':
            return 'No se seleccionó ningún archivo', 400
        
        # Guardar archivo
        filename = secure_filename(archivo.filename)
        archivo_path = os.path.join(base_dir, 'uploads', filename)
        os.makedirs(os.path.dirname(archivo_path), exist_ok=True)
        archivo.save(archivo_path)
        
        return render_template_string(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Archivo Procesado - FANGIO TELECOM</title>
            <style>
                body {{
                    background-image: url('/static/images/earth-background.jpg');
                    background-size: cover;
                    background-attachment: fixed;
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: white;
                }}
                .container {{
                    max-width: 800px;
                    margin: 0 auto;
                    background: rgba(0, 0, 0, 0.8);
                    padding: 20px;
                    border-radius: 10px;
                }}
                .logo {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .logo img {{
                    max-width: 200px;
                    height: auto;
                }}
                .btn {{
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                }}
                .btn:hover {{
                    background-color: #45a049;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
                </div>
                <h1>Archivo Procesado</h1>
                <a href="/" class="btn">← Volver al Inicio</a>
                <p>Archivo: {filename}</p>
                <p>Descripción: {descripcion}</p>
                <p>Archivo guardado exitosamente.</p>
            </div>
        </body>
        </html>
        """)
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/descargar_diseno_solucion')
def descargar_diseno_solucion():
    try:
        def limpiar_nombre_archivo(nombre):
            # Eliminar caracteres problemáticos para nombres de archivo
            caracteres_invalidos = r'[<>:"/\\|?*]'
            nombre_limpio = re.sub(caracteres_invalidos, '_', str(nombre))
            return nombre_limpio

        @after_this_request
        def eliminar_archivos_temporales(response):
            try:
                # Limpiar archivos temporales si es necesario
                pass
            except Exception as e:
                print(f"Error limpiando archivos: {e}")
            return response

        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear archivo Excel con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diseño de Solución"
        
        # Llenar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar archivo
        output_path = os.path.join(base_dir, 'site_survey', 'diseno_solucion.xlsx')
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name='diseno_solucion.xlsx')
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/seleccion_tipo_llenado')
def seleccion_tipo_llenado():
    return render_template_string(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Selección de Tipo de Llenado - FANGIO TELECOM</title>
        <style>
            body {{
                background-image: url('/static/images/earth-background.jpg');
                background-size: cover;
                background-attachment: fixed;
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: white;
            }}
            .container {{
                max-width: 800px;
                margin: 0 auto;
                background: rgba(0, 0, 0, 0.8);
                padding: 20px;
                border-radius: 10px;
            }}
            .logo {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .logo img {{
                max-width: 200px;
                height: auto;
            }}
            .btn {{
                background-color: #4CAF50;
                color: white;
                padding: 15px 30px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
                margin: 10px;
                font-size: 16px;
                width: 200px;
                text-align: center;
            }}
            .btn:hover {{
                background-color: #45a049;
            }}
            .btn-container {{
                text-align: center;
                margin-top: 30px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logo">
                <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
            </div>
            <h1>Selecciona el Tipo de Llenado</h1>
            <div class="btn-container">
                <a href="/site_survey" class="btn">Site Survey</a>
                <a href="/diseno_solucion" class="btn">Diseño de Solución</a>
            </div>
            <div style="text-align: center; margin-top: 20px;">
                <a href="/" class="btn" style="width: auto;">← Volver al Inicio</a>
            </div>
        </div>
    </body>
    </html>
    """)

@app.route('/redirigir_tipo_llenado', methods=['POST'])
def redirigir_tipo_llenado():
    tipo_llenado = request.form.get('tipo_llenado')
    if tipo_llenado == 'site_survey':
        return redirect('/site_survey')
    elif tipo_llenado == 'diseno_solucion':
        return redirect('/diseno_solucion')
    else:
        return redirect('/')

# Funciones auxiliares
def normaliza_texto(texto):
    if pd.isna(texto):
        return ""
    texto_str = str(texto).lower()
    texto_str = unicodedata.normalize('NFD', texto_str)
    texto_str = ''.join(c for c in texto_str if not unicodedata.combining(c))
    return texto_str

if __name__ == "__main__":
    try:
        print("Iniciando servidor Flask...")
        print("El servidor estará disponible en:")
        print("  - Local: http://127.0.0.1:5000")
        print("  - Red: http://0.0.0.0:5000")
        print("Presiona Ctrl+C para detener el servidor")
        app.run(debug=True, use_reloader=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
    except KeyboardInterrupt:
        print("\nServidor detenido por el usuario")
    except Exception as e:
        print(f"Error iniciando servidor: {e}") 