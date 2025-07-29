#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FANGIO TELECOM - Versión para Heroku/Cloud
Sistema de Llenado Automático para equipo distribuido
"""

import os
import sys
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Para servidor sin GUI
import numpy as np
from datetime import datetime
import zipfile
import io
import base64
from docx import Document
import tempfile
import shutil

# Configuración de la aplicación
app = Flask(__name__)
app.secret_key = 'fangio_telecom_2024'

# Configuración de rutas relativas
base_dir = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(base_dir, 'uploads')
TEMPLATE_PATH = os.path.join(base_dir, 'site_survey', 'EJEMPLO SS VACIO.xlsx')
OUTPUT_FOLDER = os.path.join(base_dir, 'Temp')

# Crear carpetas si no existen
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(os.path.join(base_dir, 'site_survey'), exist_ok=True)

# Configuración de archivos permitidos
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'xlsx', 'kmz'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def limpiar_archivos_temporales():
    """Limpia archivos temporales después de cada solicitud"""
    try:
        # Limpiar archivos en Temp
        for archivo in os.listdir(OUTPUT_FOLDER):
            archivo_path = os.path.join(OUTPUT_FOLDER, archivo)
            if os.path.isfile(archivo_path):
                os.remove(archivo_path)
    except Exception as e:
        print(f"Error limpiando archivos temporales: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/seleccion_tipo')
def seleccion_tipo():
    return render_template('seleccion_tipo_llenado.html')

@app.route('/site_survey', methods=['GET', 'POST'])
def site_survey():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            archivos = request.files.getlist('archivos')
            
            # Guardar archivos subidos
            archivos_guardados = []
            for archivo in archivos:
                if archivo and allowed_file(archivo.filename):
                    filename = secure_filename(archivo.filename)
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    archivo.save(filepath)
                    archivos_guardados.append(filepath)
            
            # Crear archivo Excel con datos
            output_filename = f"site_survey_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            # Crear nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Site Survey"
            
            # Llenar datos básicos
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Agregar información de archivos
            ws[f'A{row}'] = "Archivos subidos:"
            row += 1
            for archivo in archivos_guardados:
                ws[f'A{row}'] = os.path.basename(archivo)
                row += 1
            
            # Guardar archivo
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name=output_filename)
            
        except Exception as e:
            flash(f'Error en llenado site_survey: {str(e)}', 'error')
            return redirect(url_for('site_survey'))
    
    return render_template('site_survey_checkboxes.html')

@app.route('/diseno_solucion', methods=['GET', 'POST'])
def diseno_solucion():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            archivos = request.files.getlist('archivos')
            
            # Guardar archivos subidos
            archivos_guardados = []
            for archivo in archivos:
                if archivo and allowed_file(archivo.filename):
                    filename = secure_filename(archivo.filename)
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    archivo.save(filepath)
                    archivos_guardados.append(filepath)
            
            # Crear archivo Excel con datos
            output_filename = f"diseno_solucion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            # Crear nuevo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Diseño de Solución"
            
            # Llenar datos básicos
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Agregar información de archivos
            ws[f'A{row}'] = "Archivos subidos:"
            row += 1
            for archivo in archivos_guardados:
                ws[f'A{row}'] = os.path.basename(archivo)
                row += 1
            
            # Guardar archivo
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name=output_filename)
            
        except Exception as e:
            flash(f'Error en llenado diseño de solución: {str(e)}', 'error')
            return redirect(url_for('diseno_solucion'))
    
    return render_template('formulario_archivos.html')

@app.route('/status')
def status():
    return jsonify({
        'status': 'online',
        'service': 'FANGIO TELECOM',
        'version': '2.0-cloud',
        'timestamp': datetime.now().isoformat()
    })

@app.route('/info')
def info():
    return jsonify({
        'app_name': 'FANGIO TELECOM',
        'description': 'Sistema de Llenado Automático para equipo distribuido',
        'features': [
            'Site Survey',
            'Diseño de Solución',
            'Subida de archivos',
            'Generación de Excel'
        ],
        'upload_folder': UPLOAD_FOLDER,
        'template_path': TEMPLATE_PATH,
        'output_folder': OUTPUT_FOLDER
    })

@app.after_request
def after_request(response):
    """Limpia archivos temporales después de cada respuesta"""
    limpiar_archivos_temporales()
    return response

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False) 