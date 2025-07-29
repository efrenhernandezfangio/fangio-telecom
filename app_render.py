import os
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask, request, send_file, render_template_string, redirect, url_for, after_this_request
from werkzeug.utils import secure_filename
import sys
import re
import unicodedata

def normaliza_na(valor):
    if isinstance(valor, str) and valor.strip().lower() == "n/a":
        return "N/A"
    elif pd.isna(valor):
        return "N/A"
    elif valor == "" or (isinstance(valor, str) and valor.strip() == ""):
        return "N/A"
    return valor

# Obtener el directorio base de la aplicación
base_dir = os.path.dirname(os.path.abspath(__file__))
print(f"Directorio base: {base_dir}")

# Cargar HTML principal
try:
    with open('llenado-automatico.html', encoding='utf-8') as f:
        html_form = f.read()
    print("✅ HTML cargado exitosamente")
except Exception as e:
    print(f"⚠️ Error cargando HTML: {e}")
    html_form = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>FANGIO TELECOM - Sistema de Llenado Automático</title>
        <style>
            body {
                background-image: url('/static/images/earth-background.jpg');
                background-size: cover;
                background-attachment: fixed;
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: white;
            }
            .container {
                max-width: 800px;
                margin: 0 auto;
                background: rgba(0, 0, 0, 0.8);
                padding: 20px;
                border-radius: 10px;
            }
            .logo {
                text-align: center;
                margin-bottom: 20px;
            }
            .logo img {
                max-width: 200px;
                height: auto;
            }
            .btn {
                background-color: #4CAF50;
                color: white;
                padding: 15px 30px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
                margin: 10px;
                font-size: 16px;
                width: 200px;
                text-align: center;
            }
            .btn:hover {
                background-color: #45a049;
            }
            .btn-container {
                text-align: center;
                margin-top: 30px;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logo">
                <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
            </div>
            <h1>Sistema de Llenado Automático</h1>
            <div class="btn-container">
                <a href="/site_survey" class="btn">Site Survey</a>
                <a href="/diseno_solucion" class="btn">Diseño de Solución</a>
            </div>
        </div>
    </body>
    </html>
    """

app = Flask(__name__)

# Configuración
UPLOAD_FOLDER = os.path.join(base_dir, 'site_survey')
GOOGLE_SHEETS_CSV_URL = 'https://docs.google.com/spreadsheets/d/1sfOY1Y3dNVCOT8zyCMzpgARv-R_jRE-S/export?format=csv'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template_string(html_form)

@app.route('/site_survey', methods=['GET'])
def site_survey():
    try:
        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear tabla HTML
        tabla_html = df.to_html(classes='table table-striped', index=False)
        
        return render_template_string(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Site Survey - FANGIO TELECOM</title>
            <style>
                body {{
                    background-image: url('/static/images/earth-background.jpg');
                    background-size: cover;
                    background-attachment: fixed;
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 20px;
                    color: white;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: rgba(0, 0, 0, 0.8);
                    padding: 20px;
                    border-radius: 10px;
                }}
                .logo {{
                    text-align: center;
                    margin-bottom: 20px;
                }}
                .logo img {{
                    max-width: 200px;
                    height: auto;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    background: rgba(255, 255, 255, 0.9);
                    color: black;
                }}
                th, td {{
                    padding: 8px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #4CAF50;
                    color: white;
                }}
                .btn {{
                    background-color: #4CAF50;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                    margin: 5px;
                }}
                .btn:hover {{
                    background-color: #45a049;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
                </div>
                <h1>Base de Datos - Site Survey</h1>
                <a href="/" class="btn">← Volver al Inicio</a>
                <a href="/descargar_site_survey" class="btn">📥 Descargar Excel</a>
                {tabla_html}
            </div>
        </body>
        </html>
        """)
        
    except Exception as e:
        return f"Error al cargar datos: {str(e)}", 500

@app.route('/descargar_site_survey')
def descargar_site_survey():
    try:
        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear archivo Excel con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Site Survey"
        
        # Llenar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar archivo
        output_path = os.path.join(base_dir, 'site_survey', 'site_survey_completo.xlsx')
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name='site_survey_completo.xlsx')
        
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/diseno_solucion', methods=['GET', 'POST'])
def diseno_solucion():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            datos = request.form.to_dict()
            
            # Crear archivo Excel con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Diseño de Solución"
            
            # Llenar datos en el Excel
            row = 1
            for key, value in datos.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            # Guardar archivo
            output_path = os.path.join(base_dir, 'site_survey', 'diseno_solucion.xlsx')
            wb.save(output_path)
            
            return send_file(output_path, as_attachment=True, download_name='diseno_solucion.xlsx')
            
        except Exception as e:
            return f"Error: {str(e)}", 500
    
    # Mostrar formulario de diseño de solución
    return render_template_string(f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Diseño de Solución - FANGIO TELECOM</title>
        <style>
            body {{
                background-image: url('/static/images/earth-background.jpg');
                background-size: cover;
                background-attachment: fixed;
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: white;
            }}
            .container {{
                max-width: 800px;
                margin: 0 auto;
                background: rgba(0, 0, 0, 0.8);
                padding: 20px;
                border-radius: 10px;
            }}
            .logo {{
                text-align: center;
                margin-bottom: 20px;
            }}
            .logo img {{
                max-width: 200px;
                height: auto;
            }}
            .form-group {{
                margin-bottom: 15px;
            }}
            label {{
                display: block;
                margin-bottom: 5px;
                font-weight: bold;
            }}
            input[type="text"], textarea {{
                width: 100%;
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                box-sizing: border-box;
            }}
            .btn {{
                background-color: #4CAF50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                text-decoration: none;
                display: inline-block;
                margin: 5px;
            }}
            .btn:hover {{
                background-color: #45a049;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logo">
                <img src="/static/images/fangio-logo.PNG" alt="FANGIO TELECOM">
            </div>
            <h1>Diseño de Solución</h1>
            <a href="/" class="btn">← Volver al Inicio</a>
            <form method="post">
                <div class="form-group">
                    <label for="proyecto">Nombre del Proyecto:</label>
                    <input type="text" id="proyecto" name="proyecto" required>
                </div>
                <div class="form-group">
                    <label for="cliente">Cliente:</label>
                    <input type="text" id="cliente" name="cliente" required>
                </div>
                <div class="form-group">
                    <label for="descripcion">Descripción:</label>
                    <textarea id="descripcion" name="descripcion" rows="4" required></textarea>
                </div>
                <div class="form-group">
                    <label for="fecha">Fecha:</label>
                    <input type="date" id="fecha" name="fecha" required>
                </div>
                <button type="submit" class="btn">📥 Generar Excel</button>
            </form>
        </div>
    </body>
    </html>
    """)

@app.route('/descargar_diseno_solucion')
def descargar_diseno_solucion():
    try:
        # Leer datos de Google Sheets
        df = pd.read_csv(GOOGLE_SHEETS_CSV_URL)
        
        # Crear archivo Excel con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diseño de Solución"
        
        # Llenar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Guardar archivo
        output_path = os.path.join(base_dir, 'site_survey', 'diseno_solucion.xlsx')
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name='diseno_solucion.xlsx')
        
    except Exception as e:
        return f"Error: {str(e)}", 500

# Funciones auxiliares
def normaliza_texto(texto):
    if pd.isna(texto):
        return ""
    texto_str = str(texto).lower()
    texto_str = unicodedata.normalize('NFD', texto_str)
    texto_str = ''.join(c for c in texto_str if not unicodedata.combining(c))
    return texto_str

if __name__ == "__main__":
    try:
        print("🚀 Iniciando servidor Flask...")
        print("📍 El servidor estará disponible en:")
        print("   - Local: http://127.0.0.1:5000")
        print("   - Red: http://0.0.0.0:5000")
        print("⏹️  Presiona Ctrl+C para detener el servidor")
        
        port = int(os.environ.get('PORT', 5000))
        app.run(debug=False, host='0.0.0.0', port=port)
        
    except KeyboardInterrupt:
        print("\n⏹️ Servidor detenido por el usuario")
    except Exception as e:
        print(f"❌ Error iniciando servidor: {e}") 